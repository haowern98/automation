"""
GSN VS AD Data Extractor

This processor extracts GSN VS AD comparison data from the Weekly Report Excel file.
It finds the correct worksheet based on year and extracts data for a specific date range.
"""
import os
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from src.utils.logger import write_log


class GSNvsADExtractor:
    """Class to extract GSN VS AD data from Weekly Report Excel file"""
    
    def __init__(self, excel_file_path=None):
        """
        Initialize with Excel file path
        
        Args:
            excel_file_path (str): Path to the Excel file
        """
        # If no path is provided, get from settings
        if not excel_file_path:
            try:
                from src.gui.settings_dialog import get_settings
                settings = get_settings()
                configured_path = settings.get('file_paths', 'weekly_report_file_path', '')
                
                if configured_path and os.path.exists(configured_path):
                    self.excel_file_path = configured_path
                else:
                    # Fallback to default hardcoded path
                    self.excel_file_path = self._get_default_path()
            except Exception:
                # If settings can't be loaded, use default hardcoded path
                self.excel_file_path = self._get_default_path()
        else:
            self.excel_file_path = excel_file_path
    
    def _get_default_path(self):
        """Get the default hardcoded path as fallback"""
        user_profile = os.environ.get('USERPROFILE', '')
        return os.path.join(
            user_profile, 
            'DPDHL', 
            'SM Team - SG - AD EDS, MFA, GSN VS AD, GSN VS ER Weekly Report', 
            'Weekly Report 2025.xlsx'
        )
    
    def extract_date_components(self, date_range_str):
        """
        Extract year from a date range string
        
        Args:
            date_range_str (str): Date range string (e.g., '1-3 August 2025')
            
        Returns:
            str: Year (e.g., '2025')
        """
        # Match different date range patterns to extract year
        patterns = [
            r'\d+-\d+\s+[A-Za-z]+\s+(\d{4})',  # "1-3 August 2025"
            r'\d+\s+[A-Za-z]+\s+-\s+\d+\s+[A-Za-z]+\s+(\d{4})',  # "1 Aug - 3 Sep 2025"
            r'(\d{4})'  # Just find any 4-digit year
        ]
        
        for pattern in patterns:
            match = re.search(pattern, date_range_str)
            if match:
                return match.group(1)
        
        # Default to current year if no match
        return str(datetime.datetime.now().year)
    
    def determine_worksheet_name(self, date_range_str):
        """
        Determine the GSN VS AD worksheet name based on date range
        
        Args:
            date_range_str (str): Date range string (e.g., '1-3 August 2025')
            
        Returns:
            str: Worksheet name (e.g., 'GSN VS AD 2025')
        """
        year = self.extract_date_components(date_range_str)
        return f"GSN VS AD {year}"
    
    def determine_target_row_text(self, date_range_str):
        """
        Determine the target row text to search for
        
        Args:
            date_range_str (str): Date range string (e.g., '13-17 February 2025')
            
        Returns:
            list: List of target row texts (e.g., ['13-17 February 2025 GSN VS AD', '13-17 Feb 2025 GSN VS AD'])
        """
        full_month_name = date_range_str
        abbreviated_month_name = re.sub(r'(\w+)\s+(\d{4})', lambda m: m.group(1)[:3] + ' ' + m.group(2), date_range_str)
        return [f"{full_month_name} GSN VS AD", f"{abbreviated_month_name} GSN VS AD"]
        
    def extract_gsn_vs_ad_data(self, date_range_str):
        """
        Extract GSN VS AD data for the given date range
        Uses logic similar to the TypeScript version but returns list format
        
        Args:
            date_range_str (str): Date range string (e.g., '13-17 February 2025')
            
        Returns:
            tuple: (success: bool, data: list, error_message: str)
        """
        try:
            write_log(f"=== GSN VS AD EXTRACTION START ===", "YELLOW")
            write_log(f"Input date range: '{date_range_str}'", "YELLOW")
            
            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                error_msg = f"Excel file not found: {self.excel_file_path}"
                write_log(error_msg, "RED")
                return False, [], error_msg
            
            # Determine worksheet name and target row
            worksheet_name = self.determine_worksheet_name(date_range_str)
            target_row_texts = self.determine_target_row_text(date_range_str)
            
            write_log(f"Worksheet: '{worksheet_name}'", "CYAN")
            write_log(f"Looking for: '{target_row_texts}'", "CYAN")
            
            # Load workbook with openpyxl
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
            
            # Check if worksheet exists
            if worksheet_name not in workbook.sheetnames:
                error_msg = f"Worksheet '{worksheet_name}' not found. Available: {workbook.sheetnames}"
                write_log(error_msg, "RED")
                workbook.close()
                return False, [], error_msg
            
            worksheet = workbook[worksheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            write_log(f"Worksheet loaded: {max_row} rows, {max_col} columns", "GREEN")
            
            # Find the starting row that contains the date range
            start_row_index = None
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, min(7, max_col + 1)):  # Check first 6 columns
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and any(target_row_text in str(cell_value) for target_row_text in target_row_texts):
                        start_row_index = row_idx
                        write_log(f"Found target at row {start_row_index}, col {col_idx}", "GREEN")
                        break
                if start_row_index:
                    break
            
            if start_row_index is None:
                error_msg = f"Target row text '{target_row_texts}' not found"
                write_log(error_msg, "RED")
                workbook.close()
                return False, [], error_msg
            
            # Extract data starting from the found row
            data = []
            extracted_rows_count = 0
            
            write_log(f"Starting extraction from row {start_row_index}...", "CYAN")
            
            for row_idx in range(start_row_index, max_row + 1):
                # Check if all 6 columns are empty (stopping condition)
                all_columns_empty = True
                contains_date_range = False
                
                row_data = []
                for col_idx in range(1, 7):  # 6 columns for GSN VS AD
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    
                    if cell_value is not None and str(cell_value).strip():
                        all_columns_empty = False
                        
                        # Check if this cell contains a date range (stopping condition)
                        if col_idx == 1 and not any(target_row_text in str(cell_value).strip() for target_row_text in target_row_texts):
                            # Check if this looks like another date range
                            if 'GSN VS AD' in str(cell_value) and not any(target_row_text in str(cell_value).strip() for target_row_text in target_row_texts):
                                contains_date_range = True
                                write_log(f"Found next date range at row {row_idx}: '{str(cell_value).strip()}'", "CYAN")
                        
                        # Add cell to row data
                        if pd.isna(cell_value) or cell_value is None:
                            row_data.append({'value': ''})
                        else:
                            row_data.append({'value': str(cell_value).strip()})
                
                # Stopping conditions (similar to TypeScript logic)
                if all_columns_empty and not any(target_row_text in str(worksheet.cell(row=row_idx, column=c).value or '') for target_row_text in target_row_texts for c in range(1, 7)):
                    write_log(f"Stopping at row {row_idx}: all columns empty", "YELLOW")
                    break
                
                if contains_date_range and row_idx > start_row_index:
                    write_log(f"Stopping at row {row_idx}: found next date range", "YELLOW")
                    break
                
                # Add row to data if it has any content
                if any(cell['value'] for cell in row_data):
                    data.append(row_data)
                    extracted_rows_count += 1
                    
                    # Debug: Show first few rows
                    if extracted_rows_count <= 5:
                        preview = " | ".join([cell['value'] for cell in row_data[:3]])
                        write_log(f"Row {extracted_rows_count}: {preview}...", "WHITE")
                
                # Safety limit to prevent infinite extraction
                if extracted_rows_count >= 50:
                    write_log(f"Safety limit reached: extracted {extracted_rows_count} rows", "YELLOW")
                    break
            
            workbook.close()
            
            write_log(f"=== GSN VS AD EXTRACTION SUCCESS: {len(data)} rows ===", "GREEN")
            return True, data, ""
            
        except Exception as e:
            error_msg = f"GSN VS AD extraction failed: {str(e)}"
            write_log(error_msg, "RED")
            import traceback
            write_log(traceback.format_exc(), "RED")
            return False, [], error_msg
        
def main():
    """Test function for the GSN VS AD extractor"""
    try:
        print("\nGSN VS AD Data Extractor Test")
        print("============================")
        
        # Create the extractor
        extractor = GSNvsADExtractor()
        
        # Test with a sample date range
        date_range_str = input("Enter date range to extract (e.g., '29-30 May 2025'): ").strip()
        
        print(f"\nExtracting GSN VS AD data for: {date_range_str}...")
        success, data, error_msg = extractor.extract_gsn_vs_ad_data(date_range_str)
        
        if success:
            print(f"\nSuccess! Extracted {len(data)} rows of data")
            
            # Show first few rows
            print("\nFirst 3 rows of extracted data:")
            for i, row in enumerate(data[:3]):
                row_text = " | ".join([cell['value'] for cell in row])
                print(f"Row {i+1}: {row_text}")
                
        else:
            print(f"\nError: {error_msg}")
            
    except Exception as e:
        print(f"\nError in test: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()