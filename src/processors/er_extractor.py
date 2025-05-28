"""
ER Data Extractor

This processor extracts ER data from the Weekly Report Excel file.
It finds the correct worksheet based on year and extracts data for a specific date range.
"""
import os
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from src.utils.logger import write_log


class ERExtractor:
    """Class to extract ER data from Weekly Report Excel file"""
    
    def __init__(self, excel_file_path=None):
        """
        Initialize with Excel file path
        
        Args:
            excel_file_path (str): Path to the Excel file
        """
        # If no path is provided, get from settings
        if not excel_file_path:
            try:
                from src.gui.settings_dialog import get_settings
                settings = get_settings()
                configured_path = settings.get('file_paths', 'weekly_report_file_path', '')
                
                if configured_path and os.path.exists(configured_path):
                    self.excel_file_path = configured_path
                else:
                    # Fallback to default hardcoded path
                    self.excel_file_path = self._get_default_path()
            except Exception:
                # If settings can't be loaded, use default hardcoded path
                self.excel_file_path = self._get_default_path()
        else:
            self.excel_file_path = excel_file_path
    
    def _get_default_path(self):
        """Get the default hardcoded path as fallback"""
        user_profile = os.environ.get('USERPROFILE', '')
        return os.path.join(
            user_profile, 
            'DPDHL', 
            'SM Team - SG - AD EDS, MFA, GSN VS AD, GSN VS ER Weekly Report', 
            'Weekly Report 2025.xlsx'
        )
    
    def extract_year_from_date_range(self, date_range_str):
        """
        Extract year from a date range string
        
        Args:
            date_range_str (str): Date range string (e.g., '2-3 June 2025')
            
        Returns:
            str: Year (e.g., '2025')
        """
        # Match different date range patterns to extract year
        patterns = [
            r'\d+-\d+\s+[A-Za-z]+\s+(\d{4})',  # "2-3 June 2025"
            r'\d+\s+[A-Za-z]+\s+-\s+\d+\s+[A-Za-z]+\s+(\d{4})',  # "2 Jun - 3 Jul 2025"
            r'(\d{4})'  # Just find any 4-digit year
        ]
        
        for pattern in patterns:
            match = re.search(pattern, date_range_str)
            if match:
                return match.group(1)
        
        # Default to current year if no match
        return str(datetime.datetime.now().year)
    
    def format_date_for_search(self, date_range_str):
        """
        Format date range for searching in the ER worksheet
        Converts full month names to abbreviated versions
        
        Args:
            date_range_str (str): Date range string (e.g., '12-13 June 2025')
            
        Returns:
            str: Formatted date for search (e.g., '12-13 Jun 2025')
        """
        # Convert full month names to abbreviated versions
        month_abbreviations = {
            'January': 'Jan', 'February': 'Feb', 'March': 'Mar', 'April': 'Apr',
            'May': 'May', 'June': 'Jun', 'July': 'Jul', 'August': 'Aug',
            'September': 'Sep', 'October': 'Oct', 'November': 'Nov', 'December': 'Dec'
        }
        
        formatted_date = date_range_str
        for full_month, abbrev_month in month_abbreviations.items():
            formatted_date = formatted_date.replace(full_month, abbrev_month)
        
        return formatted_date
    
    def determine_worksheet_name(self, date_range_str):
        """
        Determine the ER worksheet name based on date range
        
        Args:
            date_range_str (str): Date range string (e.g., '2-3 June 2025')
            
        Returns:
            str: Worksheet name (e.g., 'ER 2025')
        """
        year = self.extract_year_from_date_range(date_range_str)
        return f"ER {year}"
    
    def get_cell_formatting(self, cell):
        """
        Extract formatting information from a cell
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            dict: Dictionary containing formatting information
        """
        try:
            # Get background color
            bg_color = "#FFFFFF"  # Default white
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                try:
                    rgb_value = cell.fill.start_color.rgb
                    if hasattr(rgb_value, '__len__') and len(rgb_value) == 8:  # ARGB format
                        bg_color = f"#{rgb_value[2:]}"  # Remove alpha channel
                    elif hasattr(rgb_value, '__len__') and len(rgb_value) == 6:  # RGB format
                        bg_color = f"#{rgb_value}"
                    else:
                        # Handle RGB object
                        bg_color = f"#{rgb_value:06X}" if isinstance(rgb_value, int) else "#FFFFFF"
                except (TypeError, ValueError):
                    bg_color = "#FFFFFF"
            
            # Get font color
            font_color = "#000000"  # Default black
            if cell.font and cell.font.color and cell.font.color.rgb:
                try:
                    rgb_value = cell.font.color.rgb
                    if hasattr(rgb_value, '__len__') and len(rgb_value) == 8:  # ARGB format
                        font_color = f"#{rgb_value[2:]}"  # Remove alpha channel
                    elif hasattr(rgb_value, '__len__') and len(rgb_value) == 6:  # RGB format
                        font_color = f"#{rgb_value}"
                    else:
                        # Handle RGB object
                        font_color = f"#{rgb_value:06X}" if isinstance(rgb_value, int) else "#000000"
                except (TypeError, ValueError):
                    font_color = "#000000"
            
            # Get bold status
            is_bold = cell.font.bold if cell.font and cell.font.bold else False
            
            return {
                'cell_colour': bg_color,
                'font_colour': font_color,
                'isBolded': 'bold' if is_bold else 'normal'
            }
        except Exception as e:
            # Silently fall back to defaults
            return {
                'cell_colour': '#FFFFFF',
                'font_colour': '#000000',
                'isBolded': 'normal'
            }
        
    def extract_er_data(self, date_range_str):
        """
        Extract ER data for the given date range
        Following the TypeScript logic provided
        
        Args:
            date_range_str (str): Date range string (e.g., '12-13 June 2025')
            
        Returns:
            tuple: (success: bool, data: list, error_message: str)
        """
        try:
            write_log(f"=== ER EXTRACTION START ===", "YELLOW")
            write_log(f"Input date range: '{date_range_str}'", "YELLOW")
            
            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                error_msg = f"Excel file not found: {self.excel_file_path}"
                write_log(error_msg, "RED")
                return False, [], error_msg
            
            # Determine worksheet name and search date format
            worksheet_name = self.determine_worksheet_name(date_range_str)
            search_date_range = self.format_date_for_search(date_range_str)
            
            write_log(f"Looking for worksheet: '{worksheet_name}'", "CYAN")
            write_log(f"Searching for date range: '{search_date_range}'", "CYAN")
            
            # Load workbook with openpyxl (keep formatting)
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)
            
            # Check if worksheet exists
            if worksheet_name not in workbook.sheetnames:
                error_msg = f"Worksheet '{worksheet_name}' not found. Available: {workbook.sheetnames}"
                write_log(error_msg, "RED")
                workbook.close()
                return False, [], error_msg
            
            worksheet = workbook[worksheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            write_log(f"Worksheet loaded: {max_row} rows, {max_col} columns", "GREEN")
            
            # Get the used range values (equivalent to TypeScript rows)
            rows = []
            for row_idx in range(1, max_row + 1):
                row_values = []
                for col_idx in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is None:
                        row_values.append("")
                    else:
                        row_values.append(cell_value)
                rows.append(row_values)
            
            # Determine the starting row index based on the search_date_range
            start_row_index = None
            for row_idx, row in enumerate(rows):
                # Check if any cell in the row contains the search date range
                if any(search_date_range in str(cell) for cell in row):
                    start_row_index = row_idx
                    write_log(f"Found date range '{search_date_range}' at row {start_row_index + 1}", "GREEN")
                    break
            
            # If dateRange is not found, return empty result
            if start_row_index is None:
                error_msg = f"Date range '{search_date_range}' not found in worksheet"
                write_log(error_msg, "RED")
                workbook.close()
                return False, [], error_msg
            
            # Extract data starting from the found row
            body = []
            extracted_rows_count = 0
            
            write_log(f"Starting extraction from row {start_row_index + 1}...", "CYAN")
            write_log("Will stop when encountering #AEAAAA color OR end of used range", "CYAN")
            
            # Loop through each row starting from the startRowIndex (INCLUDE the date range header)
            for row_index in range(start_row_index, len(rows)):
                row = rows[row_index]
                
                # Check if this row contains another date range (stopping condition)
                # Skip the starting row itself
                if row_index > start_row_index:
                    # Look for date range patterns like "X-Y Month YYYY" or "X-Y Mon YYYY"
                    date_pattern = r'\d+-\d+\s+[A-Za-z]+\s+\d{4}'
                    for cell in row:
                        if cell and re.search(date_pattern, str(cell)):
                            # Make sure it's different from our search date range
                            if str(cell).strip() != search_date_range:
                                write_log(f"Stopping at row {row_index + 1}: found another date range '{str(cell)}'", "YELLOW")
                                workbook.close()
                                write_log(f"=== ER EXTRACTION SUCCESS: {len(body)} rows ===", "GREEN")
                                return True, body, ""
                
                # Check if ANY of the first 3 columns has #AEAAAA background color (stopping condition)
                found_aeaaaa_color = False
                
                # Check first 3 columns for the stopping condition color
                for col_index in range(min(3, len(row))):
                    cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
                    cell_formatting = self.get_cell_formatting(cell)
                    
                    # Debug: Log cell colors for first few rows only
                    if extracted_rows_count < 5:
                        write_log(f"Row {row_index + 1}, Col {col_index + 1}: Color = {cell_formatting['cell_colour']}, Value = '{str(row[col_index]) if col_index < len(row) else 'N/A'}'", "CYAN")
                    
                    # Check for #AEAAAA color (case insensitive)
                    if cell_formatting['cell_colour'].upper() == "#AEAAAA":
                        found_aeaaaa_color = True
                        write_log(f"Found #AEAAAA color at row {row_index + 1}, column {col_index + 1}", "YELLOW")
                        break
                        
                    # Also check for other potential gray colors that might be the stopping condition
                    gray_colors = ["#AEAAAA", "#AEAAAE", "#AEAAA", "#EFEFEF", "#F2F2F2", "#E0E0E0", "#D3D3D3"]
                    if cell_formatting['cell_colour'].upper() in [color.upper() for color in gray_colors]:
                        write_log(f"Found potential gray stopping color {cell_formatting['cell_colour']} at row {row_index + 1}, column {col_index + 1}", "YELLOW")
                        found_aeaaaa_color = True
                        break
                
                # Stopping condition: found #AEAAAA color in any of the first 3 columns
                if found_aeaaaa_color:
                    write_log(f"Stopping at row {row_index + 1}: found #AEAAAA color", "YELLOW")
                    break
                
                # Check if we've reached the end of used range (all first 3 cells empty)
                if all(not str(cell).strip() for cell in row[:3]):
                    write_log(f"Stopping at row {row_index + 1}: reached end of used range", "YELLOW")
                    break
                
                # Create row data for the first 3 columns (Column1, Column2, Column3)
                # Create row data for the first 3 columns (Column1, Column2, Column3)
                row_data = {}

                # Special handling for the first row (date range header)
                if row_index == start_row_index:
                    # First row should be merged across 3 columns with #AEAAAA background
                    row_data["Column1"] = {
                        "cell content": search_date_range,
                        "cell colour": "#AEAAAA",
                        "font colour": "#000000",
                        "isBolded": "bold",
                        "colspan": 3  # Indicate this should span 3 columns
                    }
                    row_data["Column2"] = {
                        "cell content": "",
                        "cell colour": "#AEAAAA", 
                        "font colour": "#000000",
                        "isBolded": "normal",
                        "merged": True  # Indicate this is part of merged cell
                    }
                    row_data["Column3"] = {
                        "cell content": "",
                        "cell colour": "#AEAAAA",
                        "font colour": "#000000", 
                        "isBolded": "normal",
                        "merged": True  # Indicate this is part of merged cell
                    }
                else:
                    # Normal processing for other rows
                    for col_index in range(3):  # Process first 3 columns
                        col_name = f"Column{col_index + 1}"
                        
                        # Get cell value
                        if col_index < len(row):
                            cell_value = row[col_index]
                            if cell_value == "" or cell_value is None:
                                cell_content = "<br>"
                            elif isinstance(cell_value, (int, float)):
                                cell_content = str(cell_value)
                            else:
                                cell_content = str(cell_value)
                        else:
                            cell_content = "<br>"
                        
                        # Get cell formatting
                        cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
                        formatting = self.get_cell_formatting(cell)
                        
                        # Wrap cell content with <b></b> if isBolded is "bold"
                        if formatting['isBolded'] == "bold":
                            cell_content = f"<b>{cell_content}</b>"
                        
                        row_data[col_name] = {
                            "cell content": cell_content,
                            "cell colour": formatting['cell_colour'],
                            "font colour": formatting['font_colour'],
                            "isBolded": formatting['isBolded']
                        }
                
                body.append(row_data)
                extracted_rows_count += 1
                
                # Debug: Show first few rows
                if extracted_rows_count <= 5:
                    col1_preview = row_data["Column1"]["cell content"][:30] if row_data["Column1"]["cell content"] != "<br>" else "empty"
                    col2_preview = row_data["Column2"]["cell content"][:30] if row_data["Column2"]["cell content"] != "<br>" else "empty"
                    col3_preview = row_data["Column3"]["cell content"][:30] if row_data["Column3"]["cell content"] != "<br>" else "empty"
                    write_log(f"Row {extracted_rows_count}: {col1_preview} | {col2_preview} | {col3_preview}", "WHITE")
                
                # Safety limit to prevent infinite extraction
                if extracted_rows_count >= 200:
                    write_log(f"Safety limit reached: extracted {extracted_rows_count} rows", "YELLOW")
                    break
            
            workbook.close()
            
            write_log(f"=== ER EXTRACTION SUCCESS: {len(body)} rows ===", "GREEN")
            return True, body, ""
            
        except Exception as e:
            error_msg = f"ER extraction failed: {str(e)}"
            write_log(error_msg, "RED")
            import traceback
            write_log(traceback.format_exc(), "RED")
            return False, [], error_msg

def main():
    """Test function for the ER extractor"""
    try:
        print("\nER Data Extractor Test")
        print("=====================")
        
        # Create the extractor
        extractor = ERExtractor()
        
        # Test with a sample date range
        date_range_str = input("Enter date range to extract (e.g., '2-3 June 2025'): ").strip()
        
        print(f"\nExtracting ER data for: {date_range_str}...")
        success, data, error_msg = extractor.extract_er_data(date_range_str)
        
        if success:
            print(f"\nSuccess! Extracted {len(data)} rows of data")
            
            # Show first few rows
            print("\nFirst 3 rows of extracted data:")
            for i, row in enumerate(data[:3]):
                col1 = row['Column1']['cell content']
                col2 = row['Column2']['cell content'] 
                col3 = row['Column3']['cell content']
                print(f"Row {i+1}: Col1='{col1}' | Col2='{col2}' | Col3='{col3}'")
                
        else:
            print(f"\nError: {error_msg}")
            
    except Exception as e:
        print(f"\nError in test: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()