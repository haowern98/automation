"""
GSN VS ER Data Extractor

This processor extracts GSN VS ER comparison data from the Weekly Report Excel file.
It finds the correct worksheet based on date range and extracts data for a specific date range.
"""
import os
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from src.utils.logger import write_log


class GSNvsERExtractor:
    """Class to extract GSN VS ER data from Weekly Report Excel file"""
    
    def __init__(self, excel_file_path=None):
        """
        Initialize with Excel file path
        
        Args:
            excel_file_path (str): Path to the Excel file
        """
        # If no path is provided, get from settings
        if not excel_file_path:
            try:
                from src.gui.settings_dialog import get_settings
                settings = get_settings()
                configured_path = settings.get('file_paths', 'weekly_report_file_path', '')
                
                if configured_path and os.path.exists(configured_path):
                    self.excel_file_path = configured_path
                else:
                    # Fallback to default hardcoded path
                    self.excel_file_path = self._get_default_path()
            except Exception:
                # If settings can't be loaded, use default hardcoded path
                self.excel_file_path = self._get_default_path()
        else:
            self.excel_file_path = excel_file_path
    
    def _get_default_path(self):
        """Get the default hardcoded path as fallback"""
        user_profile = os.environ.get('USERPROFILE', '')
        return os.path.join(
            user_profile, 
            'DPDHL', 
            'SM Team - SG - AD EDS, MFA, GSN VS AD, GSN VS ER Weekly Report', 
            'Weekly Report 2025.xlsx'
        )
    
    def format_date_for_worksheet_name(self, date_range_str):
        """
        Format date range for worksheet name
        
        Args:
            date_range_str (str): Date range string (e.g., '2-3 June 2025')
            
        Returns:
            str: Formatted date for worksheet (e.g., '2-3 Jun 2025')
        """
        # Convert full month names to abbreviated versions for worksheet names
        month_abbreviations = {
            'January': 'Jan', 'February': 'Feb', 'March': 'Mar', 'April': 'Apr',
            'May': 'May', 'June': 'Jun', 'July': 'Jul', 'August': 'Aug',
            'September': 'Sep', 'October': 'Oct', 'November': 'Nov', 'December': 'Dec'
        }
        
        formatted_date = date_range_str
        for full_month, abbrev_month in month_abbreviations.items():
            formatted_date = formatted_date.replace(full_month, abbrev_month)
        
        return formatted_date
    
    def determine_worksheet_name(self, date_range_str):
        """
        Determine the GSN VS ER worksheet name based on date range
        
        Args:
            date_range_str (str): Date range string (e.g., '2-3 June 2025')
            
        Returns:
            str: Worksheet name (e.g., 'GSN VS ER 2-3 Jun 2025')
        """
        formatted_date = self.format_date_for_worksheet_name(date_range_str)
        return f"GSN VS ER {formatted_date}"
        
    def get_cell_formatting(self, cell):
        """
        Extract formatting information from a cell
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            dict: Dictionary containing formatting information
        """
        try:
            # Get background color
            bg_color = "#FFFFFF"  # Default white
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                try:
                    rgb_value = cell.fill.start_color.rgb
                    if hasattr(rgb_value, '__len__') and len(rgb_value) == 8:  # ARGB format
                        bg_color = f"#{rgb_value[2:]}"  # Remove alpha channel
                    elif hasattr(rgb_value, '__len__') and len(rgb_value) == 6:  # RGB format
                        bg_color = f"#{rgb_value}"
                    else:
                        # Handle RGB object
                        bg_color = f"#{rgb_value:06X}" if isinstance(rgb_value, int) else "#FFFFFF"
                except (TypeError, ValueError):
                    bg_color = "#FFFFFF"
            
            # Get font color
            font_color = "#000000"  # Default black
            if cell.font and cell.font.color and cell.font.color.rgb:
                try:
                    rgb_value = cell.font.color.rgb
                    if hasattr(rgb_value, '__len__') and len(rgb_value) == 8:  # ARGB format
                        font_color = f"#{rgb_value[2:]}"  # Remove alpha channel
                    elif hasattr(rgb_value, '__len__') and len(rgb_value) == 6:  # RGB format
                        font_color = f"#{rgb_value}"
                    else:
                        # Handle RGB object
                        font_color = f"#{rgb_value:06X}" if isinstance(rgb_value, int) else "#000000"
                except (TypeError, ValueError):
                    font_color = "#000000"
            
            # Get bold status
            is_bold = cell.font.bold if cell.font and cell.font.bold else False
            
            return {
                'cell_colour': bg_color,
                'font_colour': font_color,
                'isBolded': 'bold' if is_bold else 'normal'
            }
        except Exception as e:
            # Silently fall back to defaults instead of logging every error
            return {
                'cell_colour': '#FFFFFF',
                'font_colour': '#000000',
                'isBolded': 'normal'
            }
            
    def extract_gsn_vs_er_data(self, date_range_str):
        """
        Extract GSN VS ER data for the given date range
        
        Args:
            date_range_str (str): Date range string (e.g., '2-3 June 2025')
            
        Returns:
            tuple: (success: bool, data: list, error_message: str)
        """
        try:
            write_log(f"=== GSN VS ER EXTRACTION START ===", "YELLOW")
            write_log(f"Input date range: '{date_range_str}'", "YELLOW")
            
            # Check if file exists
            if not os.path.exists(self.excel_file_path):
                error_msg = f"Excel file not found: {self.excel_file_path}"
                write_log(error_msg, "RED")
                return False, [], error_msg
            
            # Determine worksheet name
            worksheet_name = self.determine_worksheet_name(date_range_str)
            write_log(f"Looking for worksheet: '{worksheet_name}'", "CYAN")
            
            # Load workbook with openpyxl
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=False)  # Keep formatting
            
            # Check if worksheet exists
            if worksheet_name not in workbook.sheetnames:
                error_msg = f"Worksheet '{worksheet_name}' not found. Available: {workbook.sheetnames}"
                write_log(error_msg, "RED")
                workbook.close()
                return False, [], error_msg
            
            worksheet = workbook[worksheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            write_log(f"Worksheet loaded: {max_row} rows, {max_col} columns", "GREEN")
            
            # Find the starting row that contains the date range
            start_row_index = None
            for row_idx in range(1, min(20, max_row + 1)):  # Check first 20 rows
                cell_d_value = worksheet.cell(row=row_idx, column=4).value  # Column D
                if cell_d_value and "In GSN but not in ER" in str(cell_d_value):
                    start_row_index = row_idx
                    write_log(f"Found 'In GSN but not in ER' at row {start_row_index}", "GREEN")
                    break

            if start_row_index is None:
                # Fallback: start from row 1 if we can't find the header
                start_row_index = 1
                write_log("Could not find 'In GSN but not in ER' header, starting from row 1", "YELLOW")            
            # Extract data starting from the found row - matching TypeScript logic
            data = []
            extracted_rows_count = 0

            write_log(f"Starting extraction from row {start_row_index}...", "CYAN")
            write_log("Looking for end condition: column D contains 'GSN'", "CYAN")

            # Get all values as a 2D array (like TypeScript rows)
            all_values = []
            for row_idx in range(1, max_row + 1):
                row_values = []
                for col_idx in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    row_values.append(cell_value)
                all_values.append(row_values)

            # Convert to 0-based indexing for consistency with TypeScript
            start_row_index_0 = start_row_index - 1

            # Determine the end row index based on column D (index 3) being "GSN"
            end_row_index = None
            for row_index in range(start_row_index_0, len(all_values)):
                if len(all_values[row_index]) > 3 and all_values[row_index][3] == "GSN":
                    end_row_index = row_index
                    write_log(f"Found stopping condition at row {row_index + 1}: column D contains 'GSN'", "YELLOW")
                    break

            # If endRowIndex is not found, set it to the last row
            if end_row_index is None:
                end_row_index = len(all_values) - 1
                write_log("No 'GSN' found in column D, using last row", "YELLOW")

            # Loop through each row from startRowIndex to endRowIndex (matching TypeScript)
            for row_index in range(start_row_index_0, end_row_index + 1):
                if row_index >= len(all_values):
                    break
                    
                row = all_values[row_index]
                
                # Get values from columns D and E (indices 3 and 4) - matching TypeScript
                col_d_value = row[3] if len(row) > 3 else ""
                col_e_value = row[4] if len(row) > 4 else ""
                
                # Get the actual cells for formatting (1-based indexing for openpyxl)
                cell_d = worksheet.cell(row=row_index + 1, column=4)  # Column D
                cell_e = worksheet.cell(row=row_index + 1, column=5)  # Column E
                
                # Process cell content (matching TypeScript logic)
                d_content = "<br>" if col_d_value == "" or col_d_value is None else str(col_d_value)
                e_content = "<br>" if col_e_value == "" or col_e_value is None else str(col_e_value)
                
                # Get cell formatting
                d_formatting = self.get_cell_formatting(cell_d)
                e_formatting = self.get_cell_formatting(cell_e)

                # Force white backgrounds for all GSN VS ER cells
                d_formatting['cell_colour'] = '#FFFFFF'
                e_formatting['cell_colour'] = '#FFFFFF'

                # Keep font color black for visibility
                d_formatting['font_colour'] = '#000000'
                e_formatting['font_colour'] = '#000000'
                
                # Apply bold formatting if needed (matching TypeScript logic)
                if d_formatting['isBolded'] == 'bold':
                    d_content = f"<b>{d_content}</b>"
                if e_formatting['isBolded'] == 'bold':
                    e_content = f"<b>{e_content}</b>"
                
                # Create row data structure
                row_data = [
                    {
                        'value': d_content,
                        'cell_colour': '#FFFFFF',  # Force white
                        'font_colour': '#000000',  # Force black text
                        'isBolded': d_formatting['isBolded']
                    },
                    {
                        'value': e_content,
                        'cell_colour': '#FFFFFF',  # Force white
                        'font_colour': '#000000',  # Force black text
                        'isBolded': e_formatting['isBolded']
                    }
                ]                
                # Add to data (always add, even if empty - matching TypeScript)
                data.append(row_data)
                extracted_rows_count += 1
                
                # Debug: Show first few rows
                if extracted_rows_count <= 5:
                    preview_d = d_content[:30] if d_content != "<br>" else "empty"
                    preview_e = e_content[:30] if e_content != "<br>" else "empty"
                    write_log(f"Row {extracted_rows_count}: D: {preview_d} | E: {preview_e}...", "WHITE")
                
                # Safety limit to prevent infinite extraction
                if extracted_rows_count >= 100:
                    write_log(f"Safety limit reached: extracted {extracted_rows_count} rows", "YELLOW")
                    break
            
            workbook.close()
            
            write_log(f"=== GSN VS ER EXTRACTION SUCCESS: {len(data)} rows ===", "GREEN")
            return True, data, ""
            
        except Exception as e:
            error_msg = f"GSN VS ER extraction failed: {str(e)}"
            write_log(error_msg, "RED")
            import traceback
            write_log(traceback.format_exc(), "RED")
            return False, [], error_msg


def main():
    """Test function for the GSN VS ER extractor"""
    try:
        print("\nGSN VS ER Data Extractor Test")
        print("=============================")
        
        # Create the extractor
        extractor = GSNvsERExtractor()
        
        # Test with a sample date range
        date_range_str = input("Enter date range to extract (e.g., '2-3 June 2025'): ").strip()
        
        print(f"\nExtracting GSN VS ER data for: {date_range_str}...")
        success, data, error_msg = extractor.extract_gsn_vs_er_data(date_range_str)
        
        if success:
            print(f"\nSuccess! Extracted {len(data)} rows of data")
            
            # Show first few rows
            print("\nFirst 3 rows of extracted data:")
            for i, row in enumerate(data[:3]):
                d_value = row[0]['value']
                e_value = row[1]['value']
                print(f"Row {i+1}: D='{d_value}' | E='{e_value}'")
                
        else:
            print(f"\nError: {error_msg}")
            
    except Exception as e:
        print(f"\nError in test: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()